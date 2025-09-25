import logging
import os
import os.path as osp
import torch

# from torch.cuda.amp import autocast, GradScaler
import mmcv
import time
import cv2
import numpy as np
from collections import OrderedDict, defaultdict
import re
import json

import csv
import openpyxl
from openpyxl.styles import Font

from detectron2.utils.events import EventStorage
from detectron2.checkpoint import PeriodicCheckpointer
from detectron2.evaluation import (
    CityscapesInstanceEvaluator,
    CityscapesSemSegEvaluator,
    COCOEvaluator,
    COCOPanopticEvaluator,
    DatasetEvaluators,
    LVISEvaluator,
    PascalVOCDetectionEvaluator,
    SemSegEvaluator,
)

from detectron2.data.common import AspectRatioGroupedDataset
from detectron2.data import MetadataCatalog
from pytorch_lightning.lite import LightningLite  # import LightningLite

from lib.utils.setup_logger import log_first_n
from lib.utils.utils import dprint
from lib.vis_utils.image import grid_show, vis_bbox_opencv
from lib.torch_utils.torch_utils import ModelEMA
from lib.torch_utils.misc import nan_to_num
from core.utils import solver_utils
import core.utils.my_comm as comm
from core.utils.my_checkpoint import MyCheckpointer
from core.utils.my_writer import MyCommonMetricPrinter, MyJSONWriter, MyTensorboardXWriter
from core.utils.utils import get_emb_show
from core.utils.data_utils import denormalize_image
from core.gdrn_modeling.datasets.data_loader import build_gdrn_train_loader, build_gdrn_test_loader

from .engine_utils import batch_data, get_out_coor, get_out_mask
from .gdrn_evaluator import gdrn_inference_on_dataset, GDRN_Evaluator, gdrn_save_result_of_dataset
from .gdrn_custom_evaluator import GDRN_EvaluatorCustom
import ref


logger = logging.getLogger(__name__)


class GDRN_Lite(LightningLite):
    def get_evaluator(self, cfg, dataset_name, output_folder=None):
        """Create evaluator(s) for a given dataset.

        This uses the special metadata "evaluator_type" associated with
        each builtin dataset. For your own dataset, you can simply
        create an evaluator manually in your script and do not have to
        worry about the hacky if-else logic here.
        """
        if output_folder is None:
            output_folder = osp.join(cfg.OUTPUT_DIR, "inference")
        evaluator_list = []
        evaluator_type = MetadataCatalog.get(dataset_name).evaluator_type
        if evaluator_type in ["sem_seg", "coco_panoptic_seg"]:
            evaluator_list.append(
                SemSegEvaluator(
                    dataset_name,
                    distributed=True,
                    num_classes=cfg.MODEL.SEM_SEG_HEAD.NUM_CLASSES,
                    ignore_label=cfg.MODEL.SEM_SEG_HEAD.IGNORE_VALUE,
                    output_dir=output_folder,
                )
            )
        if evaluator_type in ["coco", "coco_panoptic_seg"]:
            evaluator_list.append(COCOEvaluator(dataset_name, cfg, True, output_folder))
        if evaluator_type == "coco_panoptic_seg":
            evaluator_list.append(COCOPanopticEvaluator(dataset_name, output_folder))
        if evaluator_type == "cityscapes_instance":
            assert (
                torch.cuda.device_count() >= self.global_rank
            ), "CityscapesEvaluator currently do not work with multiple machines."
            return CityscapesInstanceEvaluator(dataset_name)
        if evaluator_type == "cityscapes_sem_seg":
            assert (
                torch.cuda.device_count() >= self.global_rank
            ), "CityscapesEvaluator currently do not work with multiple machines."
            return CityscapesSemSegEvaluator(dataset_name)
        if evaluator_type == "pascal_voc":
            return PascalVOCDetectionEvaluator(dataset_name)
        if evaluator_type == "lvis":
            return LVISEvaluator(dataset_name, cfg, True, output_folder)

        _distributed = self.world_size > 1
        dataset_meta = MetadataCatalog.get(cfg.DATASETS.TRAIN[0])
        train_obj_names = dataset_meta.objs
        if evaluator_type == "bop":
            gdrn_eval_cls = GDRN_Evaluator if cfg.VAL.get("USE_BOP", False) else GDRN_EvaluatorCustom
            return gdrn_eval_cls(
                cfg, dataset_name, distributed=_distributed, output_dir=output_folder, train_objs=train_obj_names
            )

        if len(evaluator_list) == 0:
            raise NotImplementedError(
                "no Evaluator for the dataset {} with the type {}".format(dataset_name, evaluator_type)
            )
        if len(evaluator_list) == 1:
            return evaluator_list[0]
        return DatasetEvaluators(evaluator_list)

    def get_tbx_event_writer(self, out_dir, backup=False):
        tb_logdir = osp.join(out_dir, "tb")
        mmcv.mkdir_or_exist(tb_logdir)
        if backup and self.is_global_zero:
            old_tb_logdir = osp.join(out_dir, "tb_old")
            mmcv.mkdir_or_exist(old_tb_logdir)
            os.system("mv -v {} {}".format(osp.join(tb_logdir, "events.*"), old_tb_logdir))

        tbx_event_writer = MyTensorboardXWriter(tb_logdir, backend="tensorboardX")
        return tbx_event_writer

    def do_save_results(self, cfg, model, epoch=None, iteration=None):
        model_name = osp.basename(cfg.MODEL.WEIGHTS).split(".")[0]

        dataset_meta = MetadataCatalog.get(cfg.DATASETS.TRAIN[0])
        train_obj_names = dataset_meta.objs

        for dataset_name in cfg.DATASETS.TEST:
            if epoch is not None and iteration is not None:
                save_out_dir = osp.join(cfg.OUTPUT_DIR, f"inference_epoch_{epoch}_iter_{iteration}", dataset_name)
            else:
                save_out_dir = osp.join(cfg.OUTPUT_DIR, f"inference_{model_name}", dataset_name)

            data_loader = build_gdrn_test_loader(cfg, dataset_name, train_objs=train_obj_names)
            data_loader = self.setup_dataloaders(data_loader, replace_sampler=False, move_to_device=False)

            gdrn_save_result_of_dataset(
                cfg,
                model,
                data_loader,
                output_dir=save_out_dir,
                dataset_name=dataset_name,
                train_objs=train_obj_names,
                amp_test=cfg.TEST.AMP_TEST,
            )
    
    def compute_auc_add(self, errors, max_threshold=0.1, step=0.001):
        """
        Compute AUC for ADD(-S) error up to max_threshold.

        Args:
            errors (List[float]): list of ADD or ADI errors (in meter).
            max_threshold (float): max error threshold to consider.
            step (float): interval to sample thresholds.

        Returns:
            auc (float): area under accuracy curve scaled to [0,100]
        """
        errors = np.array(errors)
        thresholds = np.arange(0, max_threshold + step, step)
        accuracies = [(errors < t).astype(np.float32).mean() for t in thresholds]
        auc = np.trapz(accuracies, thresholds) / max_threshold
        return auc * 100  # return percentage

    def do_test(self, cfg, model, epoch=None, iteration=None):
        results = OrderedDict()
        lv_to_correct_base_names = defaultdict(list)  
        lv_to_min_errors = defaultdict(dict)         
        lv_to_max_bop = defaultdict(dict) 
        lv_to_add_values = defaultdict(list)
        lv_to_bop_values = defaultdict(list)      
        lv_to_table = defaultdict(lambda: defaultdict(dict))

        # 모든 evaluator 기록
        evaluator_results = []


        model_name = osp.basename(cfg.MODEL.WEIGHTS).split(".")[0]
        for dataset_name in cfg.DATASETS.TEST:
            match = re.match(r"ss6d_(.+?)_(lv\d+)[a-z\d]*_d\d+_te", dataset_name)
            if not match:
                print(f"[do_test] Invalid dataset_name format: {dataset_name}")
                continue
            obj_name = match.group(1)
            lv = match.group(2)
            if epoch is not None and iteration is not None:
                eval_out_dir = osp.join(cfg.OUTPUT_DIR, f"inference_epoch_{epoch}_iter_{iteration}", dataset_name)
            else:
                eval_out_dir = osp.join(cfg.OUTPUT_DIR, f"inference_{model_name}", dataset_name)

            evaluator = self.get_evaluator(cfg, dataset_name, eval_out_dir)
            evaluator.lite_self = self
            data_loader = build_gdrn_test_loader(cfg, dataset_name, train_objs=evaluator.train_objs)
            data_loader = self.setup_dataloaders(data_loader, replace_sampler=False, move_to_device=False)
            results_i = gdrn_inference_on_dataset(cfg, model, data_loader, evaluator, amp_test=cfg.TEST.AMP_TEST)
            results[dataset_name] = results_i
            evaluator_results.append((dataset_name, lv, evaluator))
        

        for dataset_name, lv, evaluator in evaluator_results:
            errors = evaluator.errors
            per_image_errors = evaluator.per_image_errors
            """
            if "ae" in dataset_name.lower():
                for base_name, testset_errors in per_image_errors.items():
                    if dataset_name not in testset_errors:
                        continue
                    ad = testset_errors[dataset_name]["ad"]
                    re_val = testset_errors[dataset_name]["re"]
                    te_val = testset_errors[dataset_name]["te"]
                    norm_add = testset_errors[dataset_name]["norm_add"]
                    bop = testset_errors[dataset_name]["bop"]

                    if base_name not in lv_to_min_errors[lv]:
                        lv_to_min_errors[lv][base_name] = {
                            "min_add": ad,
                            "testset": dataset_name,
                            "bop": bop,
                            "re": re_val,
                            "te": te_val,
                            "norm_add": norm_add,
                        }
            """        

            if "ae" not in dataset_name.lower():
                # part 1. matched images
                for obj_name in errors:
                    for fname in errors[obj_name]["ans"]:
                        lv_to_correct_base_names[lv].append(f"{lv}/{fname}")

                # part 2. min error
                for base_name, testset_errors in per_image_errors.items():
                    if dataset_name not in testset_errors:
                        continue
                    ad = testset_errors[dataset_name]["ad"]
                    re_val = testset_errors[dataset_name]["re"]
                    te_val = testset_errors[dataset_name]["te"]
                    norm_add = testset_errors[dataset_name]["norm_add"]
                    bop = testset_errors[dataset_name]["bop"]

                    if base_name not in lv_to_min_errors[lv] or ad < lv_to_min_errors[lv][base_name]["min_add"]:
                        lv_to_min_errors[lv][base_name] = {
                            "min_add": ad,
                            "testset": dataset_name,
                            "bop": bop,
                            "re": re_val,
                            "te": te_val,
                            "norm_add": norm_add,
                        }

                    if base_name not in lv_to_max_bop[lv] or bop > lv_to_max_bop[lv][base_name]["max_bop"]:
                        lv_to_max_bop[lv][base_name] = {
                            "max_bop": bop,
                            "testset": dataset_name,
                            "add": ad,
                            "re": re_val,
                            "te": te_val,
                            "norm_add": norm_add,
                        }

            # part 3. metric table (AE도 포함)
            for row in evaluator.big_tab:
                metric = row[0]
                value = row[-1]  # Avg
                lv_to_table[lv][metric][dataset_name] = value
            
        for lv, base_dict in lv_to_min_errors.items():
            for base_name, data in base_dict.items():
                lv_to_add_values[lv].append(data["min_add"])

        for lv, base_dict in lv_to_max_bop.items():
            for base_name, data in base_dict.items():
                lv_to_bop_values[lv].append(data["max_bop"])


        # --- Excel 저장 ---
        excel_path = osp.join(cfg.OUTPUT_DIR, "eval_excel", f"{obj_name}.xlsx")
        os.makedirs(osp.dirname(excel_path), exist_ok=True)
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "all_lv_summary"

        cur_row = 1

        for lv, metric_dict in sorted(lv_to_table.items()):
            ws_summary.cell(row=cur_row, column=1, value=f"## {lv}")
            ws_summary.cell(row=cur_row, column=1).font = Font(bold=True)
            cur_row += 1

            metrics = sorted(metric_dict.keys())
            datasets = list(dict.fromkeys(
                ds for m in metric_dict.values() for ds in m.keys()
            ))

            ws_summary.cell(row=cur_row, column=1, value="metric")
            for col_idx, ds in enumerate(datasets, start=2):
                ws_summary.cell(row=cur_row, column=col_idx, value=ds)
            cur_row += 1

            for metric in metrics:
                ws_summary.cell(row=cur_row, column=1, value=metric)
                for col_idx, ds in enumerate(datasets, start=2):
                    val_str = metric_dict[metric].get(ds, "")
                    try:
                        val_float = float(val_str)
                        ws_summary.cell(row=cur_row, column=col_idx, value=val_float)
                    except ValueError:
                        ws_summary.cell(row=cur_row, column=col_idx, value=val_str)
                cur_row += 1

            cur_row += 2

        # --- 전체 평균 AUC 계산 ---
        all_auc_vals = []

        for lv in lv_to_min_errors:
            for base_name, data in lv_to_min_errors[lv].items():
                norm_add_val = data.get("norm_add", None)
                if norm_add_val is not None:
                    auc_val = self.compute_auc_add([norm_add_val])
                    all_auc_vals.append(auc_val)

        overall_avg_auc = sum(all_auc_vals) / len(all_auc_vals) if all_auc_vals else None

        # --- all_lv_summary 시트 마지막에 덧붙이기 ---
        ws_summary.append([])
        ws_summary.append(["oracle_auc", round(overall_avg_auc, 3) if overall_avg_auc is not None else None])


        for lv in sorted(lv_to_min_errors.keys()):
            # add_results 시트
            sheet_add = wb.create_sheet(title=f"add_results_{lv}")
            image_results = lv_to_min_errors[lv]
            avg_add = sum(d["min_add"] for d in image_results.values()) / len(image_results) if image_results else None
            sheet_add.append(["avg_add", round(avg_add, 4) if avg_add is not None else None])
            sheet_add.append([])
            sheet_add.append(["image_name", "min_add", "testset", "re", "te"])
            for image_name, data in image_results.items():
                sheet_add.append([
                    image_name,
                    data["min_add"],
                    data["testset"],
                    data["re"],
                    data["te"]
                ])

            # bop_results 시트
            sheet_bop = wb.create_sheet(title=f"bop_results_{lv}")
            image_results = lv_to_max_bop[lv]
            avg_bop = sum(d["max_bop"] for d in image_results.values()) / len(image_results) if image_results else None
            sheet_bop.append(["avg_bop", round(avg_bop, 4) if avg_bop is not None else None])
            sheet_bop.append([])
            sheet_bop.append(["image_name", "min_bop", "testset", "add", "re", "te"])
            for image_name, data in image_results.items():
                sheet_bop.append([
                    image_name,
                    data["max_bop"],
                    data["testset"],
                    data["add"],
                    data["re"],
                    data["te"]
                ])  

            # match_imgs 시트 (add_5 기준)
            sheet_match = wb.create_sheet(title=f"match_imgs_{lv}")
            matched_base_names = sorted(set(osp.basename(p) for p in lv_to_correct_base_names[lv]))
            sheet_match.append(["num_correct", len(matched_base_names)])
            sheet_match.append([])
            sheet_match.append(["matched_images"])
            for name in matched_base_names:
                sheet_match.append([name])

            # 이미지별 testset 간 add 값 시트
            base_names = sorted(lv_to_min_errors[lv].keys())
            all_testsets = sorted({
                data["testset"] for data in lv_to_min_errors[lv].values()
            })

            sheet_add_matrix = wb.create_sheet(title=f"add_per_image_{lv}")
            sheet_add_matrix.append(["image_name"] + all_testsets)

            for base_name in base_names:
                row = [base_name]
                for testset in all_testsets:
                    if lv_to_min_errors[lv][base_name]["testset"] == testset:
                        row.append(lv_to_min_errors[lv][base_name]["min_add"])
                    else:
                        row.append("")
                sheet_add_matrix.append(row)

        wb.save(excel_path)

        if len(results) == 1:
            results = list(results.values())[0]
        return results

    def do_train(self, cfg, args, model, optimizer, renderer=None, resume=False):
        model.train()

        # some basic settings =========================
        dataset_meta = MetadataCatalog.get(cfg.DATASETS.TRAIN[0])
        data_ref = ref.__dict__[dataset_meta.ref_key]
        obj_names = dataset_meta.objs

        # load data ===================================
        train_dset_names = cfg.DATASETS.TRAIN
        data_loader = build_gdrn_train_loader(cfg, train_dset_names)

        data_loader_iter = iter(data_loader)

        # load 2nd train dataloader if needed
        train_2_dset_names = cfg.DATASETS.get("TRAIN2", ())
        train_2_ratio = cfg.DATASETS.get("TRAIN2_RATIO", 0.0)
        if train_2_ratio > 0.0 and len(train_2_dset_names) > 0:
            data_loader_2 = build_gdrn_train_loader(cfg, train_2_dset_names)
            data_loader_2_iter = iter(data_loader_2)
        else:
            data_loader_2 = None
            data_loader_2_iter = None

        images_per_batch = cfg.SOLVER.IMS_PER_BATCH
        if isinstance(data_loader, AspectRatioGroupedDataset):
            dataset_len = len(data_loader.dataset.dataset)
            if data_loader_2 is not None and cfg.DATASETS.DATA_LEN_WITH_TRAIN2:
                logger.info("Calculate dataset length with TRAIN2")
                dataset_len += len(data_loader_2.dataset.dataset)
            iters_per_epoch = dataset_len // images_per_batch
        else:
            dataset_len = len(data_loader.dataset)
            if data_loader_2 is not None and cfg.DATASETS.DATA_LEN_WITH_TRAIN2:
                logger.info("Calculate dataset length with TRAIN2")
                dataset_len += len(data_loader_2.dataset)
            iters_per_epoch = dataset_len // images_per_batch
        max_iter = cfg.SOLVER.TOTAL_EPOCHS * iters_per_epoch
        dprint("images_per_batch: ", images_per_batch)
        dprint("dataset length: ", dataset_len)
        dprint("iters per epoch: ", iters_per_epoch)
        dprint("total iters: ", max_iter)

        data_loader = self.setup_dataloaders(data_loader, replace_sampler=False, move_to_device=False)
        if data_loader_2 is not None:
            data_loader_2 = self.setup_dataloaders(data_loader_2, replace_sampler=False, move_to_device=False)

        bs_ref = cfg.SOLVER.get("REFERENCE_BS", 64)  # nominal batch size =========================
        accumulate_iter = max(round(bs_ref / cfg.SOLVER.IMS_PER_BATCH), 1)  # accumulate loss before optimizing
        # NOTE: update lr every accumulate_iter
        scheduler = solver_utils.build_lr_scheduler(cfg, optimizer, total_iters=max_iter // accumulate_iter)

        # resume or load model ===================================
        extra_ckpt_dict = dict(
            optimizer=optimizer,
            scheduler=scheduler,
        )
        if hasattr(self._precision_plugin, "scaler"):
            extra_ckpt_dict["gradscaler"] = self._precision_plugin.scaler
        checkpointer = MyCheckpointer(
            model,
            cfg.OUTPUT_DIR,
            save_to_disk=self.is_global_zero,
            prefix_to_remove="_module.",
            **extra_ckpt_dict,
        )
        # start_iter = checkpointer.resume_or_load(cfg.MODEL.WEIGHTS, resume=resume).get("iteration", -1) + 1
        
        # load ckpt
        ckpt = checkpointer.resume_or_load(cfg.MODEL.WEIGHTS, resume=resume)
        start_iter = ckpt.get("iteration", -1) + 1

        # 꼭 필요!
        if "scheduler" in ckpt:
            scheduler.load_state_dict(ckpt["scheduler"])

        # Exponential moving average (NOTE: initialize ema after loading weights) ========================
        if self.is_global_zero and cfg.MODEL.EMA.ENABLED:
            ema = ModelEMA(model, **cfg.MODEL.EMA.INIT_CFG)
            ema.updates = start_iter // accumulate_iter
            # save the ema model
            checkpointer.model = ema.ema.module if hasattr(ema.ema, "module") else ema.ema
        else:
            ema = None

        if cfg.SOLVER.CHECKPOINT_BY_EPOCH:
            ckpt_period = cfg.SOLVER.CHECKPOINT_PERIOD * iters_per_epoch
        else:
            ckpt_period = cfg.SOLVER.CHECKPOINT_PERIOD
        periodic_checkpointer = PeriodicCheckpointer(
            checkpointer, ckpt_period, max_iter=max_iter, max_to_keep=cfg.SOLVER.MAX_TO_KEEP
        )

        # build writers ==============================================
        tbx_event_writer = self.get_tbx_event_writer(cfg.OUTPUT_DIR, backup=not cfg.get("RESUME", False))
        tbx_writer = tbx_event_writer._writer  # NOTE: we want to write some non-scalar data
        writers = (
            [MyCommonMetricPrinter(max_iter), MyJSONWriter(osp.join(cfg.OUTPUT_DIR, "metrics.json")), tbx_event_writer]
            if self.is_global_zero
            else []
        )

        # compared to "train_net.py", we do not support accurate timing and
        # precise BN here, because they are not trivial to implement
        logger.info("Starting training from iteration {}".format(start_iter))
        iter_time = None
        last_evaled_epoch = -1
        with EventStorage(start_iter) as storage:
            optimizer.zero_grad(set_to_none=True)
            for iteration in range(start_iter, max_iter):
                storage.iter = iteration
                epoch = iteration // iters_per_epoch + 1  # epoch start from 1
                storage.put_scalar("epoch", epoch, smoothing_hint=False)

                if np.random.rand() < train_2_ratio:
                    data = next(data_loader_2_iter)
                else:
                    data = next(data_loader_iter)

                if iter_time is not None:
                    storage.put_scalar("time", time.perf_counter() - iter_time)
                iter_time = time.perf_counter()

                # if cfg.TRAIN.VIS:
                #     vis_train_data(data, obj_names, cfg)

                # forward ============================================================
                batch = batch_data(cfg, data, renderer=renderer)
                if cfg.INPUT.WITH_DEPTH:
                    inp = torch.cat([batch["roi_img"], batch["roi_depth"]], dim=1)
                else:
                    inp = batch["roi_img"]

                out_dict, loss_dict = model(
                    inp,
                    gt_xyz=batch.get("roi_xyz", None),
                    gt_xyz_bin=batch.get("roi_xyz_bin", None),
                    gt_mask_trunc=batch["roi_mask_trunc"],
                    gt_mask_visib=batch["roi_mask_visib"],
                    gt_mask_full=batch.get("roi_mask_full", None),
                    gt_mask_obj=batch["roi_mask_obj"],
                    gt_region=batch.get("roi_region", None),
                    gt_ego_rot=batch.get("ego_rot", None),
                    gt_trans=batch.get("trans", None),
                    gt_trans_ratio=batch["roi_trans_ratio"],
                    gt_points=batch.get("roi_points", None),
                    sym_infos=batch.get("sym_info", None),
                    roi_classes=batch["roi_cls"],
                    roi_cams=batch["roi_cam"],
                    roi_whs=batch["roi_wh"],
                    roi_centers=batch["roi_center"],
                    resize_ratios=batch["resize_ratio"],
                    roi_coord_2d=batch.get("roi_coord_2d", None),
                    roi_coord_2d_rel=batch.get("roi_coord_2d_rel", None),
                    roi_extents=batch.get("roi_extent", None),
                    do_loss=True,
                )
                losses = sum(loss_dict.values())
                assert torch.isfinite(losses).all(), loss_dict

                if self.is_global_zero:
                    log_first_n(logging.INFO, "iteration {} forward finished.".format(iteration), n=2)

                loss_dict_reduced = {k: v.item() for k, v in comm.reduce_dict(loss_dict).items()}
                losses_reduced = sum(loss for loss in loss_dict_reduced.values())
                if self.is_global_zero:
                    storage.put_scalars(total_loss=losses_reduced, **loss_dict_reduced)

                # backward & optimize ======================================================
                self.backward(losses)
                # optimize
                # set nan grads to 0
                if cfg.SOLVER.SET_NAN_GRAD_TO_ZERO:
                    for param in model.parameters():
                        if param.grad is not None:
                            nan_to_num(param.grad, nan=0, posinf=1e5, neginf=-1e5, out=param.grad)
                optimizer.step()

                if self.is_global_zero:
                    log_first_n(logging.INFO, "iteration {} backward finished.".format(iteration), n=2)
                if iteration % accumulate_iter == 0:
                    optimizer.zero_grad(set_to_none=True)
                    if ema is not None:
                        ema.update(model)
                    storage.put_scalar("lr", optimizer.param_groups[0]["lr"], smoothing_hint=False)
                    scheduler.step()

                if (
                    cfg.TEST.EVAL_PERIOD > 0
                    and epoch != last_evaled_epoch
                    and (epoch % cfg.TEST.EVAL_PERIOD == 0)
                    and iteration != max_iter - 1
                ):
                    last_evaled_epoch = epoch
                    if ema is not None:
                        ema.update_attr(model)
                        self.do_test(
                            cfg,
                            model=ema.ema.module if hasattr(ema.ema, "module") else ema.ema,
                            epoch=epoch,
                            iteration=iteration,
                        )
                    else:
                        self.do_test(cfg, model, epoch=epoch, iteration=iteration)
                    # Compared to "train_net.py", the test results are not dumped to EventStorage
                    self.barrier()

                if iteration - start_iter > 5 and (
                    (iteration + 1) % cfg.TRAIN.PRINT_FREQ == 0 or iteration == max_iter - 1 or iteration < 100
                ):
                    for writer in writers:
                        writer.write()
                    # visualize some images ========================================
                    if cfg.TRAIN.VIS_IMG:
                        with torch.no_grad():
                            vis_i = 0
                            roi_img_vis = batch["roi_img"][vis_i].cpu().numpy()
                            roi_img_vis = denormalize_image(roi_img_vis, cfg).transpose(1, 2, 0).astype("uint8")
                            tbx_writer.add_image("input_image", roi_img_vis, iteration)

                            out_coor_x = out_dict["coor_x"].detach()
                            out_coor_y = out_dict["coor_y"].detach()
                            out_coor_z = out_dict["coor_z"].detach()
                            out_xyz = get_out_coor(cfg, out_coor_x, out_coor_y, out_coor_z)

                            out_xyz_vis = out_xyz[vis_i].cpu().numpy().transpose(1, 2, 0)
                            out_xyz_vis = get_emb_show(out_xyz_vis)
                            tbx_writer.add_image("out_xyz", out_xyz_vis, iteration)

                            gt_xyz_vis = batch["roi_xyz"][vis_i].cpu().numpy().transpose(1, 2, 0)
                            gt_xyz_vis = get_emb_show(gt_xyz_vis)
                            tbx_writer.add_image("gt_xyz", gt_xyz_vis, iteration)

                            out_mask = out_dict["mask"].detach()
                            out_mask = get_out_mask(cfg, out_mask)
                            out_mask_vis = out_mask[vis_i, 0].cpu().numpy()
                            tbx_writer.add_image("out_mask", out_mask_vis, iteration)

                            gt_mask_vis = batch["roi_mask"][vis_i].detach().cpu().numpy()
                            tbx_writer.add_image("gt_mask", gt_mask_vis, iteration)
                if (iteration + 1) % periodic_checkpointer.period == 0 or (
                    periodic_checkpointer.max_iter is not None and (iteration + 1) >= periodic_checkpointer.max_iter
                ):
                    if hasattr(optimizer, "consolidate_state_dict"):  # for ddp_sharded
                        optimizer.consolidate_state_dict()
                periodic_checkpointer.step(iteration, epoch=epoch)


def vis_train_data(data, obj_names, cfg):
    for i, d in enumerate(data):
        # if i >= 1:
        #     continue
        full_img = mmcv.imread(d["file_name"], "color")
        # if "000009/rgb/000047.png" not in d["file_name"]:
        #     continue
        print(d["file_name"])
        im_H, im_W = full_img.shape[:2]
        roi_cls = d["roi_cls"]
        if roi_cls not in [0]:
            continue
        bbox_center = d["bbox_center"]
        scale = d["scale"]
        x1 = max(min(bbox_center[0] - scale / 2, im_W), 0)
        x2 = max(min(bbox_center[0] + scale / 2, im_W), 0)
        y1 = max(min(bbox_center[1] - scale / 2, im_H), 0)
        y2 = max(min(bbox_center[1] + scale / 2, im_H), 0)
        full_img_vis = vis_bbox_opencv(full_img, np.array([x1, y1, x2, y2]), fmt="xyxy")

        bbox_ori = d["bbox"]
        full_img_bbox = vis_bbox_opencv(full_img, bbox_ori, fmt="xyxy")
        obj_name = obj_names[roi_cls]

        roi_img = d["roi_img"].numpy()
        roi_img = denormalize_image(roi_img, cfg).transpose(1, 2, 0).astype("uint8")

        roi_mask_trunc = d["roi_mask_trunc"].numpy().astype("bool")
        roi_mask_visib = d["roi_mask_visib"].numpy().astype("bool")
        roi_mask_obj = d["roi_mask_obj"].numpy().astype("bool")

        kernel = np.ones((3, 3), np.uint8)
        erode_mask_obj = cv2.erode(roi_mask_obj.astype("uint8"), kernel, iterations=1)

        roi_xyz = d["roi_xyz"].numpy().transpose(1, 2, 0)
        roi_xyz_show = get_emb_show(roi_xyz) * erode_mask_obj[:, :, None].astype("float32")

        coord2d = d["roi_coord_2d"].numpy().transpose(1, 2, 0)
        roi_h, roi_w = coord2d.shape[:2]
        zeros_1 = np.zeros((roi_h, roi_w, 1), dtype="float32")
        coord2d_3 = np.concatenate([zeros_1, get_emb_show(coord2d)], axis=2)

        # yapf: disable
        vis_imgs = [
            full_img_vis[:, :, [2, 1, 0]], full_img_bbox[:, :, [2, 1, 0]], roi_img[:, :, [2, 1, 0]],
            roi_mask_trunc * erode_mask_obj, roi_mask_visib*erode_mask_obj, roi_mask_obj*erode_mask_obj,
            roi_xyz_show,
            coord2d_3,
            coord2d[:, :, 0], coord2d[:, :, 1]
        ]
        titles = [
            "full_img", "ori_bbox", f"roi_img({obj_name})",
            "roi_mask_trunc",  "roi_mask_visib", "roi_mask_obj",
            "roi_xyz",
            "roi_coord2d",
            "roi_coord2d_x", "roi_coord2d_y"
        ]
        row = 3
        col = 4
        if "roi_region" in d:
            roi_region = d["roi_region"].numpy()  # (bh, bw)
            roi_region_3 = np.zeros((roi_h, roi_w, 3), dtype="float32")
            for region_id in range(256):
                # if region_id == 0:
                #     continue
                if region_id in roi_region:
                    for _c in range(3):
                        roi_region_3[:, :, _c][roi_region == region_id] = roi_xyz_show[:, :, _c][roi_region == region_id].mean()
            roi_region_3 = roi_region_3  * erode_mask_obj[:, :, None].astype("float32")
            vis_imgs.append(roi_region_3)
            titles.append("roi_region")
        if len(vis_imgs) > row * col:
            col += 1
        for _im, _name in zip(vis_imgs, titles):
            save_path = osp.join(cfg.OUTPUT_DIR, "vis", _name+'.png')
            mmcv.mkdir_or_exist(osp.dirname(save_path))
            if _im.shape[-1] == 3:
                _im = _im[:, :, [2,1,0]]
            if _im.max() < 1.1:
                _im = (_im * 255).astype("uint8")
            print(save_path)
            mmcv.imwrite(_im, save_path)

        grid_show(vis_imgs, titles, row=row, col=col)

        # yapf: enable
